# app.py
import os
import sqlite3
import secrets
import io
import json
import urllib.request
import re
from datetime import datetime
from pathlib import Path
from flask import (
    Flask, request, jsonify, g, render_template_string, 
    send_from_directory, session, redirect, url_for, send_file
)
import openpyxl

# ===== 설정 및 경로 =====
APP_DIR = Path(__file__).resolve().parent if "__file__" in locals() else Path.cwd()
DB_PATH = APP_DIR / "items.db"
UPLOAD_DIR = (APP_DIR / "uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = APP_DIR / "access_log.txt"
ALLOWED_IMAGE_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp"}

app = Flask(__name__)
app.secret_key = "my_secret_key_v7_final_fix"
app.config["UPLOAD_FOLDER"] = str(UPLOAD_DIR)
ADMIN_PASSWORD = "dltmdgus"

# DB 스키마
SCHEMA = """
CREATE TABLE IF NOT EXISTS items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    category TEXT NOT NULL,
    title TEXT NOT NULL,
    details TEXT,
    image TEXT,
    created_at TEXT NOT NULL
);
"""

# ===== 유틸리티 함수 =====
def get_country(ip):
    if ip in ["127.0.0.1", "localhost", "::1"]: return "Local Network"
    try:
        url = f"http://ip-api.com/json/{ip}"
        with urllib.request.urlopen(url, timeout=2) as response:
            data = json.loads(response.read().decode())
            return data.get("country", "Unknown")
    except: return "Trace Error"

def write_access_log(status="Access"):
    ip = request.headers.get('X-Forwarded-For', request.remote_addr).split(',')[0]
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    country = get_country(ip)
    path = request.path
    log_entry = f"[{now}] IP: {ip} | Country: {country} | Path: {path} | Status: {status}\n"
    with open(LOG_FILE, "a", encoding="utf-8") as f: 
        f.write(log_entry)

@app.before_request
def before_request_logging():
    if not request.path.startswith('/uploads/') and request.path != "/login":
        write_access_log("Access")

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(str(DB_PATH), check_same_thread=False)
        g.db.row_factory = sqlite3.Row
        g.db.execute(SCHEMA)
        cols = {r[1] for r in g.db.execute("PRAGMA table_info(items)").fetchall()}
        if "image" not in cols: g.db.execute("ALTER TABLE items ADD COLUMN image TEXT")
        g.db.commit()
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db: db.close()

def login_required(f):
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"): return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper

def save_image(image_file):
    if image_file and image_file.filename != "":
        ext = os.path.splitext(image_file.filename)[1].lower()
        if ext in ALLOWED_IMAGE_EXT:
            img_name = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{secrets.token_hex(4)}{ext}"
            image_file.save(UPLOAD_DIR / img_name)
            return img_name
    return None

# ===== API & Routes =====
@app.route("/")
@login_required
def index(): return render_template_string(HTML_TEMPLATE)

@app.route("/lldp")
@login_required
def lldp_page(): return render_template_string(LLDP_TEMPLATE)

@app.route("/shift")
@login_required
def shift_page(): return render_template_string(SHIFT_TEMPLATE)

@app.route("/api/ddos/summarize", methods=["POST"])
@login_required
def summarize_ddos():
    text = request.json.get("text", "")
    time_match = re.search(r"발생시간\s*:\s*\d{4}-\d{2}-\d{2}\s+(\d{2}:\d{2})", text)
    mo_match = re.search(r"대상\s*M\.O\s*:\s*(.*)", text)
    ip_match = re.search(r"대상\s*IP\s*:\s*(.*)", text)
    capa_match = re.search(r"용량\s*:\s*([\d\.]+\s*(?:Gbps|Mbps))", text, re.I)
    result_match = re.search(r"조치 결과\s*:\s*(.*)", text)
    time_val = time_match.group(1) if time_match else "00:00"
    mo_val = mo_match.group(1).strip() if mo_match else "Unknown"
    ip_val = ip_match.group(1).strip() if ip_match else "0.0.0.0"
    capa_val = capa_match.group(1).strip() if capa_match else "0 Gbps"
    result_val = result_match.group(1).strip() if result_match else "미처리"
    summary = f"({time_val}) {mo_val} / {ip_val} / {capa_val} 발생 / {result_val} / 특이사항 없음"
    return jsonify({"summary": summary})

@app.route("/login", methods=["GET", "POST"])
def login_page():
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            session["logged_in"] = True
            write_access_log("Login Success")
            return redirect(url_for("index"))
        write_access_log("Login Failed")
        return "비밀번호 오류. <a href='/login'>다시 시도</a>"
    return '<html><body style="display:flex;justify-content:center;align-items:center;height:100vh;background:#f0f2f5;margin:0;"><form method="POST" style="background:#fff;padding:40px;border-radius:15px;box-shadow:0 10px 25px rgba(0,0,0,0.1);width:300px;"><h2>MyWiki Login</h2><input type="password" name="password" style="width:100%;padding:12px;margin:15px 0;border:1px solid #ddd;border-radius:8px;" placeholder="Password" autofocus><button style="width:100%;padding:12px;background:#4f46e5;color:#fff;border:none;border-radius:8px;cursor:pointer;font-weight:bold;">Login</button></form></body></html>'

@app.route("/logout")
def logout(): session.pop("logged_in", None); return redirect(url_for("login_page"))

@app.route("/api/items", methods=["GET"])
@login_required
def get_items():
    db = get_db()
    rows = db.execute("SELECT * FROM items ORDER BY category, title").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/items", methods=["POST"])
@login_required
def add_item():
    img = save_image(request.files.get("image"))
    db = get_db()
    db.execute("INSERT INTO items (category, title, details, image, created_at) VALUES (?,?,?,?,?)", (request.form['category'], request.form['title'], request.form['details'], img, datetime.now().isoformat()))
    db.commit()
    return jsonify({"success": True})

@app.route("/api/items/<int:item_id>", methods=["PUT"])
@login_required
def update_item(item_id):
    db = get_db()
    new_img = save_image(request.files.get("image"))
    if new_img:
        old = db.execute("SELECT image FROM items WHERE id=?", (item_id,)).fetchone()[0]
        if old and (UPLOAD_DIR / old).exists(): os.remove(UPLOAD_DIR / old)
        db.execute("UPDATE items SET category=?, title=?, details=?, image=? WHERE id=?", (request.form['category'], request.form['title'], request.form['details'], new_img, item_id))
    else:
        db.execute("UPDATE items SET category=?, title=?, details=? WHERE id=?", (request.form['category'], request.form['title'], request.form['details'], item_id))
    db.commit()
    return jsonify({"success": True})

@app.route("/api/items/<int:item_id>", methods=["DELETE"])
@login_required
def delete_item(item_id):
    db = get_db()
    old = db.execute("SELECT image FROM items WHERE id=?", (item_id,)).fetchone()[0]
    if old and (UPLOAD_DIR / old).exists(): os.remove(UPLOAD_DIR / old)
    db.execute("DELETE FROM items WHERE id=?", (item_id,))
    db.commit()
    return jsonify({"success": True})

@app.route("/uploads/<path:filename>")
def serve_upload(filename): return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

@app.route("/api/excel/download")
@login_required
def excel_download():
    db = get_db()
    rows = db.execute("SELECT category, title, details, created_at FROM items ORDER BY category, title").fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["category", "title", "details", "created_at"])
    for r in rows: ws.append([r[0], r[1], r[2], r[3]])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    today = datetime.now().strftime('%Y%m%d')
    filename = f"myWIKI_{today}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/excel/upload", methods=["POST"])
@login_required
def excel_upload():
    mode = request.form.get("mode")
    file = request.files.get('file')
    if not file: return jsonify({"success": False, "error": "파일이 없습니다."}), 400
    db = get_db()
    if mode == 'reset':
        rows = db.execute("SELECT image FROM items WHERE image IS NOT NULL").fetchall()
        for r in rows:
            if r[0] and (UPLOAD_DIR / r[0]).exists(): os.remove(UPLOAD_DIR / r[0])
        db.execute("DELETE FROM items")
        db.execute("DELETE FROM sqlite_sequence WHERE name='items'")
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            db.execute("INSERT INTO items (category, title, details, created_at) VALUES (?,?,?,?)", 
                       (str(row[0] or "기타"), str(row[1]), str(row[2] or ""), datetime.now().isoformat()))
    db.commit()
    return jsonify({"success": True})

# ===== HTML Templates =====
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <title>MyWiki v6</title>
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        :root { --primary: #4f46e5; --bg: #f3f4f6; --card: #ffffff; --text: #1f2937; --border: #e5e7eb; }
        * { box-sizing: border-box; font-family: -apple-system, sans-serif; }
        body { margin: 0; background: var(--bg); color: var(--text); display: flex; flex-direction: column; height: 100vh; overflow: hidden; }
        header { background: var(--card); padding: 15px 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); flex-shrink: 0; z-index: 100; }
        .header-top { display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px; }
        #search-input { width: 100%; padding: 12px; border-radius: 10px; border: 1px solid var(--border); outline: none; }
        main { display: flex; flex: 1; overflow: hidden; }
        #list-container { width: 350px; overflow-y: auto; padding: 15px; border-right: 1px solid var(--border); background: #f8fafc; }
        #detail-container { flex: 1; overflow-y: auto; padding: 30px; background: var(--card); }
        .item-card { background: white; padding: 12px; border-radius: 8px; margin-bottom: 8px; cursor: pointer; border: 1px solid transparent; transition: 0.2s; box-shadow: 0 1px 2px rgba(0,0,0,0.05); }
        .item-card:hover { border-color: var(--primary); }
        .btn { background: var(--primary); color: white; border: none; padding: 8px 16px; border-radius: 6px; cursor: pointer; font-weight: bold; }
        .btn-outline { background: white; color: var(--primary); border: 1px solid var(--primary); }
        .modal { position: fixed; inset: 0; background: rgba(0,0,0,0.5); display: none; align-items: center; justify-content: center; z-index: 1000; padding: 20px; }
        .modal-content { background: white; width: 100%; max-width: 500px; padding: 25px; border-radius: 15px; max-height: 90vh; overflow-y: auto; }
        input, textarea { width: 100%; padding: 12px; margin: 8px 0; border: 1px solid var(--border); border-radius: 8px; }
        .detail-img { max-width: 100%; border-radius: 10px; margin: 20px 0; box-shadow: 0 4px 12px rgba(0,0,0,0.1); }
        @media (max-width: 768px) {
            #list-container { width: 100%; }
            #detail-container { display: none; position: fixed; inset: 0; z-index: 500; }
            #detail-container.mobile-show { display: block; }
        }
        .settings-dropdown { position: relative; display: inline-block; }
        .settings-content { display: none; position: absolute; right: 0; background-color: #fff; min-width: 180px; box-shadow: 0px 8px 16px 0px rgba(0,0,0,0.2); z-index: 1100; border-radius: 8px; padding: 10px 0; }
        .settings-content a { color: black; padding: 12px 16px; text-decoration: none; display: block; cursor: pointer; font-size: 14px; }
        .settings-content a:hover { background-color: #f1f1f1; }
        .show { display: block; }
    </style>
</head>
<body>
<header>
    <div class="header-top">
        <h2 style="margin:0; color:var(--primary);">MyWiki</h2>
        <div style="display:flex; gap:8px; align-items: center;">
            <button class="btn" style="background:#0ea5e9;" onclick="window.open('/shift', '_blank')">야간근무</button>
            <button class="btn" style="background:#10b981;" onclick="window.open('/lldp', '_blank')">LLDP</button>
            <button class="btn" onclick="openModal('add')">추가</button> 
            <div class="settings-dropdown">
                <button class="btn btn-outline" onclick="toggleSettings()"><i class="fas fa-cog"></i></button>
                <div id="settings-menu" class="settings-content">
                    <a onclick="location.href='/api/excel/download'"><i class="fas fa-download"></i> 엑셀파일로 다운로드</a>
                    <a onclick="triggerExcelUpload('reset')"><i class="fas fa-trash-alt"></i> 엑셀로 전체 새로등록</a>
                    <a onclick="triggerExcelUpload('append')"><i class="fas fa-plus"></i> 엑셀로 추가 업로드</a>
                    <hr style="margin: 5px 0; border: 0; border-top: 1px solid #eee;">
                    <a onclick="location.href='/logout'" style="color: #e11d48;"><i class="fas fa-sign-out-alt"></i> 로그아웃</a>
                </div>
            </div>
            <input type="file" id="excel-file-input" style="display:none;" accept=".xlsx" onchange="processExcelUpload()">
        </div>
    </div>
    <input type="text" id="search-input" placeholder="검색어를 입력하세요..." oninput="renderList()">
</header>
<main>
    <div id="list-container"></div>
    <div id="detail-container">
        <button class="btn btn-outline" style="margin-bottom:20px;" onclick="closeDetail()">← 목록보기</button>
        <div id="detail-view"></div>
    </div>
</main>
<div id="item-modal" class="modal" onclick="if(event.target==this) closeItemModal()">
    <div class="modal-content">
        <h3 id="modal-title">편집</h3>
        <input type="hidden" id="form-id"><input type="text" id="form-category" placeholder="카테고리"><input type="text" id="form-title" placeholder="제목">
        <textarea id="form-details" rows="10" placeholder="상세 내용"></textarea>
        <input type="file" id="form-image" accept="image/*">
        <div style="display:flex; gap:10px; margin-top:20px;"><button class="btn" style="flex:2" onclick="saveItem()">저장</button><button class="btn btn-outline" style="flex:1" onclick="closeItemModal()">취소</button></div>
    </div>
</div>
<script>
    let dbData = [];
    let currentUploadMode = 'append';
    async function loadData() { const r = await fetch('/api/items'); dbData = await r.json(); renderList(); }
    function toggleSettings() { document.getElementById("settings-menu").classList.toggle("show"); }
    window.onclick = function(event) { if (!event.target.matches('.btn-outline') && !event.target.matches('.fa-cog')) { var dropdowns = document.getElementsByClassName("settings-content"); for (var i = 0; i < dropdowns.length; i++) { var openDropdown = dropdowns[i]; if (openDropdown.classList.contains('show')) { openDropdown.classList.remove('show'); } } } }
    function triggerExcelUpload(mode) { currentUploadMode = mode; if(confirm(mode === 'reset' ? "기존 데이터를 모두 삭제하고 엑셀 내용으로 새로 등록하시겠습니까?" : "엑셀 내용을 기존 데이터에 추가하시겠습니까?")) { document.getElementById('excel-file-input').click(); } }
    async function processExcelUpload() { const fileInput = document.getElementById('excel-file-input'); if(!fileInput.files[0]) return; const fd = new FormData(); fd.append('file', fileInput.files[0]); fd.append('mode', currentUploadMode); const res = await fetch('/api/excel/upload', { method: 'POST', body: fd }); if(res.ok) { alert("완료!"); location.reload(); } }
    function renderList() { const q = document.getElementById('search-input').value.toLowerCase(); const container = document.getElementById('list-container'); container.innerHTML = ''; const groups = {}; dbData.filter(i => i.title.toLowerCase().includes(q) || (i.details && i.details.toLowerCase().includes(q))).forEach(i => { if(!groups[i.category]) groups[i.category] = []; groups[i.category].push(i); }); for(const cat in groups) { const gDiv = document.createElement('div'); gDiv.innerHTML = `<div style="font-weight:bold; color:#64748b; font-size:12px; margin:15px 0 5px 5px;">${cat}</div>`; groups[cat].forEach(item => { const card = document.createElement('div'); card.className = 'item-card'; card.innerText = item.title; card.onclick = () => showDetail(item.id); gDiv.appendChild(card); }); container.appendChild(gDiv); } }
    function showDetail(id) { const item = dbData.find(i => i.id === id); const img = item.image ? `<img src="/uploads/${item.image}" class="detail-img">` : ''; document.getElementById('detail-view').innerHTML = `<div style="display:flex; justify-content:space-between;"><div><span style="color:var(--primary); font-weight:bold;">#${item.category}</span><h1>${item.title}</h1></div><div style="display:flex; gap:5px;"><button class="btn btn-outline" onclick="editItem(${item.id})">수정</button><button class="btn" style="background:#fee2e2; color:#b91c1c;" onclick="deleteItem(${item.id})">삭제</button></div></div><hr>${img}<div style="white-space:pre-wrap; line-height:1.7;">${item.details || ''}</div>`; document.getElementById('detail-container').classList.add('mobile-show'); }
    function openModal(mode) { document.getElementById('item-modal').style.display='flex'; document.getElementById('modal-title').innerText='새 항목 추가'; document.getElementById('form-id').value=''; document.getElementById('form-category').value=''; document.getElementById('form-title').value=''; document.getElementById('form-details').value=''; document.getElementById('form-image').value=''; }
    function closeItemModal() { document.getElementById('item-modal').style.display='none'; }
    function closeDetail() { document.getElementById('detail-container').classList.remove('mobile-show'); }
    function editItem(id) { const item = dbData.find(i => i.id === id); document.getElementById('item-modal').style.display='flex'; document.getElementById('modal-title').innerText='항목 수정'; document.getElementById('form-id').value=item.id; document.getElementById('form-category').value=item.category; document.getElementById('form-title').value=item.title; document.getElementById('form-details').value=item.details; document.getElementById('form-image').value=''; }
    async function saveItem() { const id = document.getElementById('form-id').value; const fd = new FormData(); fd.append('category', document.getElementById('form-category').value); fd.append('title', document.getElementById('form-title').value); fd.append('details', document.getElementById('form-details').value); const img = document.getElementById('form-image').files[0]; if(img) fd.append('image', img); await fetch(id ? `/api/items/${id}` : '/api/items', { method: id ? 'PUT' : 'POST', body: fd }); closeItemModal(); loadData(); }
    async function deleteItem(id) { if(confirm('삭제?')) { await fetch(`/api/items/${id}`, { method: 'DELETE' }); loadData(); closeDetail(); } }
    window.onload = loadData;
</script>
</body>
</html>
"""

LLDP_TEMPLATE = """
<!DOCTYPE html>
<html lang="ko">
<head><meta charset="UTF-8"><title>LLDP 비교</title><style>body{font-family:sans-serif;padding:20px;background:#f3f4f6;} .grid{display:grid;grid-template-columns:1fr 1fr;gap:20px;} textarea{width:100%;height:200px;padding:10px;} .btn{width:100%;padding:15px;background:#4f46e5;color:white;border:none;cursor:pointer;margin-top:10px;}</style></head>
<body>
    <h2>LLDP 비교</h2>
    <div class="grid"><textarea id="ta" placeholder="A"></textarea><textarea id="tb" placeholder="B"></textarea></div>
    <button class="btn" onclick="compare()">비교하기</button>
    <div id="res" style="margin-top:20px;"></div>
    <script>
        async function compare(){
            const r = await fetch('/api/lldp/compare', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({text_a:document.getElementById('ta').value, text_b:document.getElementById('tb').value})});
            const d = await r.json();
            document.getElementById('res').innerHTML = `<pre>${JSON.stringify(d, null, 2)}</pre>`;
        }
    </script>
</body>
</html>
"""

SHIFT_TEMPLATE = """
<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>야간근무 v2</title>
    <style>
        :root { --primary: #4f46e5; --primary-dark: #3730a3; --slate-50: #f8fafc; --slate-100: #f1f5f9; --slate-200: #e2e8f0; --slate-700: #334155; --slate-800: #1e293b; }
        body { font-family: -apple-system, "Segoe UI", "Malgun Gothic", sans-serif; background: var(--slate-50); color: var(--slate-800); margin: 0; padding: 20px; }
        .container { max-width: 1000px; margin: 0 auto; background: white; padding: 40px; border-radius: 20px; box-shadow: 0 10px 25px rgba(0,0,0,0.05); }
        h2 { font-size: 1.5rem; font-weight: 800; color: var(--slate-800); margin-top: 0; border-bottom: 2px solid var(--slate-100); padding-bottom: 10px; }
        .section-label { font-size: 0.95rem; font-weight: 700; color: var(--slate-700); margin: 20px 0 10px 0; display: block; }
        textarea { width: 100%; height: 100px; padding: 15px; border: 2px solid var(--slate-200); border-radius: 12px; font-size: 14px; outline: none; background: var(--slate-50); }
        textarea:focus { border-color: var(--primary); background: white; }
        .worker-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 12px; margin-bottom: 20px; }
        .worker-card { background: white; padding: 12px; border-radius: 12px; border: 1px solid var(--slate-200); display: flex; align-items: center; justify-content: space-between; }
        .worker-name { font-weight: 700; font-size: 0.9rem; }
        select { padding: 6px 10px; border-radius: 8px; border: 1px solid var(--slate-200); background: white; font-size: 0.85rem; }
        .result-box { margin-top: 20px; background: var(--slate-800); border-radius: 12px; padding: 20px; color: white; }
        #rd, #summary-output { white-space: pre-wrap; font-size: 0.95rem; line-height: 1.6; font-family: monospace; }
        .btn-action { width: 100%; background: var(--primary); color: white; border: none; padding: 14px; border-radius: 10px; cursor: pointer; font-weight: 700; margin-top: 15px; }
        .btn-copy { background: #64748b; color: white; border: none; padding: 8px 16px; border-radius: 6px; cursor: pointer; font-size: 13px; margin-top: 10px; }
        hr { border: 0; border-top: 2px dashed var(--slate-200); margin: 40px 0; }
    </style>
</head>
<body>
    <div class="container">
        <h2>📅 보안팀 야간 근무 현황 정리</h2>
        <label class="section-label">1. 인원 명단 구성 (쉼표로 구분)</label>
        <textarea id="aw" oninput="gen()"></textarea>
        <label class="section-label">2. 개인별 근태 상태 지정</label>
        <div id="wl" class="worker-grid"></div>
        <label class="section-label">3. 최종 요약 결과</label>
        <div class="result-box"><div id="rd"></div></div>
        <button class="btn-action" onclick="res_copy('rd')">근무 현황 복사</button>

        <hr>

        <h2 style="color:#e11d48;">🛡️ DDoS 이벤트 요약 도구</h2>
        <label class="section-label">1. 발생 정보 입력</label>
        <textarea id="ddos-input" placeholder="DDoS 발생 정보를 붙여넣으세요..."></textarea>
        <button class="btn-action" style="background:#e11d48;" onclick="summarizeDDoS()">이벤트 요약하기</button>
        <div id="ddos-result-area" style="display:none;">
            <label class="section-label">2. 요약 결과</label>
            <div class="result-box" style="background:#f1f5f9; color:#1e293b; border: 1px solid #cbd5e1;">
                <div id="summary-output"></div>
            </div>
            <button class="btn-copy" onclick="res_copy('summary-output')">요약 결과 복사</button>
        </div>
    </div>

    <script>
        // 야간근무 로직
        const dw = "권창주, 문우석, 박정욱, 박지현, 신라성, 안미애, 안호준, 윤영화, 이슬기, 이승현(28), 이승현(22), 정준희, 최승민";
        function gen(){
            const names = document.getElementById('aw').value.split(/[,\\s\\n]+/).filter(n => n.trim());
            const list = document.getElementById('wl');
            const currentStates = {};
            document.querySelectorAll('.ws').forEach(s => currentStates[s.dataset.name] = s.value);
            list.innerHTML = names.map(n => {
                const saved = currentStates[n] || "출근";
                return `<div class="worker-card"><span class="worker-name">${n}</span><select class="ws" data-name="${n}" onchange="res()"><option ${saved==='출근'?'selected':''}>출근</option><option ${saved==='비번'?'selected':''}>비번</option><option ${saved==='야간'?'selected':''}>야간</option><option ${saved==='휴가'?'selected':''}>휴가</option><option ${saved==='재택'?'selected':''}>재택</option><option ${saved==='외근'?'selected':''}>외근</option></select></div>`;
            }).join('');
            res();
        }
        function res(){
            const now = new Date(); const cat = { "비번":[], "야간":[], "휴가":[], "재택":[], "출근":[], "외근":[] };
            document.querySelectorAll('.ws').forEach(s => cat[s.value].push(s.dataset.name));
            let r = `${now.getMonth()+1}/${now.getDate()}(${['일','월','화','수','목','금','토'][now.getDay()]}) 보안탐지대응팀 근무자 현황\\n`;
            ["비번","야간","휴가","재택","출근","외근"].forEach(k => { if(cat[k].length) r += `- ${k} : ${cat[k].join(', ')}\\n`; });
            document.getElementById('rd').innerText = r.trim();
        }

        // DDoS 로직
        async function summarizeDDoS() {
            const text = document.getElementById('ddos-input').value;
            const res = await fetch('/api/ddos/summarize', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ text: text })
            });
            const data = await res.json();
            document.getElementById('ddos-result-area').style.display = 'block';
            document.getElementById('summary-output').innerText = data.summary;
        }

        // 공통 복사 기능
        function res_copy(id) {
            const text = document.getElementById(id).innerText;
            navigator.clipboard.writeText(text).then(() => alert("복사되었습니다!"));
        }

        window.onload = () => { document.getElementById('aw').value = dw; gen(); };
    </script>
</body>
</html>
"""

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=14738, debug=True)